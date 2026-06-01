import csv
import logging

import openpyxl


def read_xlsx(filename, num_headers=1):
    """Read an Excel file and return a dict mapping sheet names to rows.

    Each sheet is parsed into a list of dictionaries using the header row(s)
    as column names. When ``num_headers >= 2``, sheets whose second row
    contains at least one ``None`` value are treated as having multi-row
    headers: the first two rows are joined with ``_`` to form combined
    column names (e.g. ``TaxonName`` + ``ASV1`` → ``TaxonName_ASV1``).
    Sheets where the second row has no ``None`` values are assumed to have
    a single header row (the second row is treated as data).

    Examples:
        >>> read_xlsx('data.xlsx')  # doctest: +SKIP
        {'Site': [{'site_name': 'Lake X', 'lat': '45.0'}], 'Samples': [...]}

    Args:
        filename (str): Path to the .xlsx file to read.
        num_headers (int): Number of header rows declared in the template.
            When >= 2, multi-row header combining is attempted per sheet.
            Defaults to 1.

    Returns:
        dict: Mapping of sheet name (str) to list of row dicts.
    """
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    except FileNotFoundError:
        logging.error(f"Excel file not found: {filename}")
        return {}
    except Exception as e:
        logging.error(f"Error opening Excel file {filename}: {e}")
        return {}

    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        if num_headers >= 2 and len(rows) >= 2 and any(v is None for v in rows[1]):
            headers = _combine_header_rows(rows[0], rows[1])
            data_start = num_headers
        else:
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_start = 1
        result[sheet_name] = [
            dict(zip(headers, row, strict=False))
            for row in rows[data_start:]
            if any(v is not None for v in row)
        ]
    wb.close()
    return result


def _combine_header_rows(row0, row1):
    """Combine two header rows into a single list of column names.

    For columns where row1 is ``None`` or empty, uses row0 alone.
    For columns where both rows have values, joins them with ``_``.

    Args:
        row0 (tuple): First header row values.
        row1 (tuple): Second header row values.

    Returns:
        list[str]: Combined column names.
    """
    headers = []
    for i, (h0, h1) in enumerate(zip(row0, row1, strict=False)):
        if h0 is None:
            headers.append(f"col_{i}")
        elif h1 is None or str(h1).strip() == "":
            headers.append(str(h0))
        else:
            headers.append(f"{h0}_{h1}")
    return headers


def read_csv(filename):
    """Read CSV file and return a structured list of dictionaries.

    Parses a CSV file and converts each row into a dictionary with column headers
    as keys.

    Examples:
        >>> read_csv('pollen_data.csv')  # doctest: +SKIP
        [{'depth': '2.5', 'quercus': '125'}, {'depth': '5.0', 'quercus': '142'}]

    Args:
        filename (str): Path to the CSV file to read.

    Returns:
        list: List of dictionaries where each dictionary represents a row,
              with column headers as keys.
    """
    with open(filename) as f:
        try:
            file_data = csv.reader(f)
            headers = next(file_data)
        except FileNotFoundError:
            logging.error(f"CSV file not found: {filename}")
            return []
        except StopIteration:
            logging.error(f"CSV file is empty: {filename}")
            return []
        except Exception as e:
            logging.error(f"Error reading CSV file {filename}: {e}")
            return []

        try:
            return [dict(zip(headers, row, strict=False)) for row in file_data]
        except Exception as e:
            logging.error(f"Error parsing CSV rows from {filename}: {e}")
            return []
