"""Tests for neotomaHelpers utilities."""

import os

import openpyxl
import pytest

import DataBUS.neotomaHelpers as nh
from tests.conftest import toy_csv, toy_yml


class TestReadCsv:
    def test_read_csv_returns_list_of_dicts(self):
        rows = nh.read_csv(toy_csv("test_sisal.csv"))
        assert isinstance(rows, list)
        assert len(rows) > 0
        assert isinstance(rows[0], dict)

    def test_read_csv_210pb(self):
        rows = nh.read_csv(toy_csv("test_210Pb.csv"))
        assert len(rows) > 0

    def test_read_csv_node(self):
        rows = nh.read_csv(toy_csv("test_node.csv"))
        assert len(rows) > 0

    def test_read_csv_eanode(self):
        rows = nh.read_csv(toy_csv("test_eanode.csv"))
        assert len(rows) > 0


class TestTemplateToDist:
    def test_template_to_dict_sisal(self):
        d = nh.template_to_dict(toy_yml("test_sisal_template.yml"))
        assert isinstance(d, dict)
        assert "metadata" in d
        assert len(d["metadata"]) > 0

    def test_template_to_dict_210pb(self):
        d = nh.template_to_dict(toy_yml("test_210pb_template.yml"))
        assert isinstance(d, dict)

    def test_template_to_dict_node(self):
        d = nh.template_to_dict(toy_yml("test_node_template.yml"))
        assert isinstance(d, dict)

    def test_template_to_dict_eanode(self):
        d = nh.template_to_dict(toy_yml("test_eanode_template.yml"))
        assert isinstance(d, dict)


class TestPullParams:
    def test_pull_params_sisal_sites(self, sisal_pair):
        csv_file, yml_dict = sisal_pair
        from DataBUS.Site import SITE_PARAMS

        result = nh.pull_params(SITE_PARAMS, yml_dict, csv_file, "ndb.sites")
        assert isinstance(result, dict)

    def test_pull_params_returns_dict(self, pb210_pair):
        csv_file, yml_dict = pb210_pair
        from DataBUS.Site import SITE_PARAMS

        result = nh.pull_params(SITE_PARAMS, yml_dict, csv_file, "ndb.sites")
        assert isinstance(result, dict)


class TestReadXlsx:
    def test_read_xlsx_returns_dict(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sites"
        ws.append(["site_name", "lat", "lon"])
        ws.append(["Lake X", "45.0", "-90.0"])
        ws2 = wb.create_sheet("Samples")
        ws2.append(["depth", "taxon"])
        ws2.append(["2.5", "Quercus"])
        ws2.append(["5.0", "Pinus"])
        path = str(tmp_path / "test.xlsx")
        wb.save(path)

        result = nh.read_xlsx(path)
        assert isinstance(result, dict)
        assert "Sites" in result
        assert "Samples" in result

    def test_read_xlsx_sheet_rows(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sites"
        ws.append(["site_name", "lat"])
        ws.append(["Lake X", "45.0"])
        path = str(tmp_path / "test.xlsx")
        wb.save(path)

        result = nh.read_xlsx(path)
        assert len(result["Sites"]) == 1
        assert result["Sites"][0]["site_name"] == "Lake X"

    def test_read_xlsx_multiple_rows(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Samples"
        ws.append(["depth", "count"])
        ws.append(["2.5", "10"])
        ws.append(["5.0", "20"])
        path = str(tmp_path / "test.xlsx")
        wb.save(path)

        result = nh.read_xlsx(path)
        assert len(result["Samples"]) == 2

    def test_read_xlsx_missing_file(self):
        result = nh.read_xlsx("nonexistent.xlsx")
        assert result == {}

    def test_read_xlsx_empty_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)

        result = nh.read_xlsx(path)
        assert result["Empty"] == []


class TestPullParamsWithSheet:
    def test_pull_params_sheet_selection(self, tmp_path):
        """pull_params picks the correct sheet when csv_file is a dict."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SiteMeta"
        ws.append(["Site.name", "Lat", "Long"])
        ws.append(["TestLake", "45.0", "-90.0"])
        path = str(tmp_path / "multi.xlsx")
        wb.save(path)

        csv_file = nh.read_xlsx(path)
        yml_dict = {
            "metadata": [
                {
                    "column": "Site.name",
                    "neotoma": "ndb.sites.sitename",
                    "sheet": "SiteMeta",
                    "rowwise": False,
                    "type": "string",
                }
            ]
        }
        result = nh.pull_params(["sitename"], yml_dict, csv_file, "ndb.sites")
        assert result.get("sitename") == "TestLake"

    def test_pull_params_missing_sheet_returns_none(self, tmp_path):
        """If the named sheet doesn't exist, pull_params returns None for the param."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OtherSheet"
        ws.append(["Site.name"])
        ws.append(["TestLake"])
        path = str(tmp_path / "wrong_sheet.xlsx")
        wb.save(path)

        csv_file = nh.read_xlsx(path)
        yml_dict = {
            "metadata": [
                {
                    "column": "Site.name",
                    "neotoma": "ndb.sites.sitename",
                    "sheet": "SiteMeta",
                    "rowwise": False,
                    "type": "string",
                }
            ]
        }
        result = nh.pull_params(["sitename"], yml_dict, csv_file, "ndb.sites")
        assert result.get("sitename") is None

    def test_pull_params_csv_list_ignores_sheet_field(self):
        """A plain list csv_file (CSV path) still works when YAML has a sheet field."""
        csv_file = [{"Site.name": "PlainCSVLake"}]
        yml_dict = {
            "metadata": [
                {
                    "column": "Site.name",
                    "neotoma": "ndb.sites.sitename",
                    "sheet": "SomSheet",
                    "rowwise": False,
                    "type": "string",
                }
            ]
        }
        result = nh.pull_params(["sitename"], yml_dict, csv_file, "ndb.sites")
        assert result.get("sitename") == "PlainCSVLake"


class TestToyData:
    """Quick sanity checks that the toy data files are readable."""

    def test_toy_sisal_readable(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_sisal.csv"))
        assert len(rows) > 0

    def test_toy_210pb_readable(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_210Pb.csv"))
        assert len(rows) > 0

    def test_toy_node_readable(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_node.csv"))
        assert len(rows) > 0

    def test_toy_eanode_readable(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_eanode.csv"))
        assert len(rows) > 0

    def test_toy_sisal_site_name_modified(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_sisal.csv"))
        assert rows[0].get("site_name") == "TestSite_SISAL"

    def test_toy_210pb_site_name_modified(self):
        from tests.conftest import toy_csv

        rows = nh.read_csv(toy_csv("test_210Pb.csv"))
        assert rows[0].get("Site.name") == "TestSite_210Pb"
